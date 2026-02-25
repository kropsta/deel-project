"""
insurance_lead_finder.py
------------------------
Pay-per-call brokerage lead scraper.
Searches Indeed, Glassdoor, ZipRecruiter, and Craigslist for insurance
agencies/call centers that are potential buyers of inbound calls and live transfers.
Outputs a professional .xlsx spreadsheet with 3 tabs + HIGH PRIORITY flagging.

Run with:
    python insurance_lead_finder.py
"""

import os
import time
import random
import hashlib
import datetime
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────

SEARCH_TERMS = [
    "health insurance agent",
    "life insurance agent",
    "medicare insurance agent",
    "ACA health insurance",
    "final expense insurance",
    "insurance call center",
    "health insurance sales",
    "medicare supplement sales",
    "warm leads insurance",
    "inbound calls insurance",
    "live transfers insurance",
    "inbound leads health insurance",
    "warm transfers medicare",
    "live transfer final expense",
]

PRIORITY_KEYWORDS = [
    "warm leads",
    "inbound calls",
    "live transfers",
    "warm transfers",
    "inbound leads",
    "calls provided",
    "leads provided",
]

CRAIGSLIST_CITIES = [
    "newyork", "losangeles", "chicago", "houston", "phoenix",
    "philadelphia", "sanantonio", "sandiego", "dallas", "austin",
    "miami", "atlanta", "seattle", "denver", "boston",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
}

MIN_DELAY = 1.5
MAX_DELAY = 3.0

# ─────────────────────────────────────────────
#  COLOURS / STYLES
# ─────────────────────────────────────────────

DARK_BLUE    = "1A2D42"
MID_BLUE     = "2A4D6E"
ROW_ALT      = "EBF4F8"
ROW_WHITE    = "FFFFFF"
GREEN_FILL   = "C6EFCE"  # HIGH PRIORITY rows
GREEN_FONT   = "276221"

thin_side = Side(border_style="thin", color="B0BEC5")
thin_border = Border(
    left=thin_side, right=thin_side,
    top=thin_side, bottom=thin_side
)

def hdr_style():
    return {
        "font": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor=DARK_BLUE),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": thin_border,
    }

def apply_style(cell, style_dict):
    for attr, val in style_dict.items():
        setattr(cell, attr, val)

# ─────────────────────────────────────────────
#  UTILITY
# ─────────────────────────────────────────────

def delay():
    s = random.uniform(MIN_DELAY, MAX_DELAY)
    time.sleep(s)

def is_high_priority(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in PRIORITY_KEYWORDS)

def get_priority(job: dict) -> str:
    combined = f"{job.get('title','')} {job.get('description','')}".lower()
    return "HIGH" if is_high_priority(combined) else "Normal"

def dedup(leads: list) -> list:
    seen = {}
    for lead in leads:
        key = hashlib.md5(lead["company"].strip().lower().encode()).hexdigest()
        if key not in seen:
            seen[key] = lead
        else:
            # keep HIGH priority if duplicate
            if lead.get("priority") == "HIGH":
                seen[key] = lead
    return list(seen.values())

def get_page(url: str, timeout: int = 10):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r.text
    except Exception as e:
        print(f"    ⚠ Request failed: {e}")
        return None

# ─────────────────────────────────────────────
#  SCRAPERS
# ─────────────────────────────────────────────

def scrape_indeed(term: str) -> list:
    results = []
    q = term.replace(" ", "+")
    url = f"https://www.indeed.com/jobs?q={q}&l=United+States"
    print(f"  [Indeed] '{term}'")
    html = get_page(url)
    delay()
    if not html:
        return results
    soup = BeautifulSoup(html, "html.parser")

    # Indeed card selectors (changes frequently; best-effort)
    cards = (
        soup.select("div.job_seen_beacon") or
        soup.select("div.resultContent") or
        soup.select("div[data-jk]")
    )
    for card in cards[:15]:
        try:
            title_el  = card.select_one("h2.jobTitle a, h2 a")
            comp_el   = card.select_one("[data-testid='company-name'], .companyName")
            loc_el    = card.select_one("[data-testid='text-location'], .companyLocation")
            desc_el   = card.select_one(".job-snippet, .underShelfFooter")

            if not title_el or not comp_el:
                continue

            title    = title_el.get_text(strip=True)
            company  = comp_el.get_text(strip=True)
            location = loc_el.get_text(strip=True) if loc_el else "US"
            desc     = desc_el.get_text(strip=True) if desc_el else ""
            href     = title_el.get("href", "")
            link     = f"https://www.indeed.com{href}" if href.startswith("/") else href

            results.append({
                "company": company, "title": title, "location": location,
                "source": "Indeed", "link": link, "search_term": term,
                "description": desc,
            })
        except Exception:
            pass
    return results


def scrape_glassdoor(term: str) -> list:
    results = []
    q = term.replace(" ", "-")
    url = f"https://www.glassdoor.com/Job/jobs.htm?sc.keyword={q.replace('-','+')}&loc=United+States"
    print(f"  [Glassdoor] '{term}'")
    html = get_page(url)
    delay()
    if not html:
        return results
    soup = BeautifulSoup(html, "html.parser")

    cards = (
        soup.select("li.react-job-listing") or
        soup.select("[data-test='jobListing']")
    )
    for card in cards[:15]:
        try:
            title_el  = card.select_one("a.jobLink, [data-test='job-link'], a[class*='jobLink']")
            comp_el   = card.select_one("[data-test='employer-name'], .job-search-card__company-name")
            loc_el    = card.select_one("[data-test='emp-location'], .job-search-card__location")

            if not title_el or not comp_el:
                continue

            title    = title_el.get_text(strip=True)
            company  = comp_el.get_text(strip=True)
            location = loc_el.get_text(strip=True) if loc_el else "US"
            href     = title_el.get("href", "")
            link     = f"https://www.glassdoor.com{href}" if href.startswith("/") else href

            results.append({
                "company": company, "title": title, "location": location,
                "source": "Glassdoor", "link": link, "search_term": term,
                "description": title,
            })
        except Exception:
            pass
    return results


def scrape_ziprecruiter(term: str) -> list:
    results = []
    q = term.replace(" ", "+")
    url = f"https://www.ziprecruiter.com/candidate/search?search={q}&location=United+States"
    print(f"  [ZipRecruiter] '{term}'")
    html = get_page(url)
    delay()
    if not html:
        return results
    soup = BeautifulSoup(html, "html.parser")

    cards = (
        soup.select("article.job_result") or
        soup.select("[class*='job_result']") or
        soup.select("li[class*='job-listing']")
    )
    for card in cards[:15]:
        try:
            title_el  = card.select_one("h2 a, .job_title a, a[class*='job_title']")
            comp_el   = card.select_one(".company_name, [class*='company']")
            loc_el    = card.select_one(".location, [class*='location']")
            desc_el   = card.select_one(".job_description, [class*='description']")

            if not title_el or not comp_el:
                continue

            title    = title_el.get_text(strip=True)
            company  = comp_el.get_text(strip=True)
            location = loc_el.get_text(strip=True) if loc_el else "US"
            desc     = desc_el.get_text(strip=True) if desc_el else ""
            href     = title_el.get("href", "")
            link     = href if href.startswith("http") else f"https://www.ziprecruiter.com{href}"

            results.append({
                "company": company, "title": title, "location": location,
                "source": "ZipRecruiter", "link": link, "search_term": term,
                "description": desc,
            })
        except Exception:
            pass
    return results


def scrape_craigslist(term: str) -> list:
    results = []
    cities = random.sample(CRAIGSLIST_CITIES, min(5, len(CRAIGSLIST_CITIES)))
    q = term.replace(" ", "+")
    for city in cities:
        url = f"https://{city}.craigslist.org/search/jjj?query={q}"
        print(f"  [Craigslist-{city}] '{term}'")
        html = get_page(url)
        delay()
        if not html:
            continue
        soup = BeautifulSoup(html, "html.parser")
        items = soup.select("li.result-row") or soup.select("[data-pid]")
        for item in items[:8]:
            try:
                title_el  = item.select_one(".result-title, a.result-title")
                if not title_el:
                    continue
                title   = title_el.get_text(strip=True)
                href    = title_el.get("href", "")
                link    = href if href.startswith("http") else f"https://{city}.craigslist.org{href}"
                company = title.split(" - ")[-1] if " - " in title else title
                results.append({
                    "company": company.strip(), "title": title, "location": city.title(),
                    "source": "Craigslist", "link": link, "search_term": term,
                    "description": title,
                })
            except Exception:
                pass
    return results


# ─────────────────────────────────────────────
#  FALLBACK SAMPLE DATA
# ─────────────────────────────────────────────

SAMPLE_DATA = [
    {"company": "National Health Advisors", "title": "Health Insurance Sales Agent (Inbound Calls Provided)", "location": "Dallas, TX", "source": "Indeed", "link": "https://www.indeed.com", "search_term": "health insurance agent", "description": "Warm leads, inbound calls provided daily. Final expense and ACA specialists welcome."},
    {"company": "MediCover Agency", "title": "Medicare Supplement Agent – Live Transfers", "location": "Phoenix, AZ", "source": "ZipRecruiter", "link": "https://www.ziprecruiter.com", "search_term": "medicare insurance agent", "description": "We provide live transfers to our agents. No cold calling required."},
    {"company": "Freedom Final Expense", "title": "Final Expense Insurance Agent – Calls Provided", "location": "Atlanta, GA", "source": "Indeed", "link": "https://www.indeed.com", "search_term": "final expense insurance", "description": "Fully vetted inbound leads provided. Work from home. Warm transfers to experienced closers."},
    {"company": "BlueSky Health Group", "title": "ACA Health Insurance Specialist", "location": "Miami, FL", "source": "Glassdoor", "link": "https://www.glassdoor.com", "search_term": "ACA health insurance", "description": "Looking for ACA specialists. Leads provided through our proprietary system."},
    {"company": "Senior Life Solutions", "title": "Medicare Advantage Sales (Remote)", "location": "Houston, TX", "source": "ZipRecruiter", "link": "https://www.ziprecruiter.com", "search_term": "medicare supplement sales", "description": "Inbound calls and warm transfers provided by our call center. Competitive commissions."},
    {"company": "Apex Insurance Group", "title": "Life Insurance Sales Agent", "location": "Chicago, IL", "source": "Indeed", "link": "https://www.indeed.com", "search_term": "life insurance agent", "description": "Proven system with warm inbound leads. Agents average 8-12 presentations per day."},
    {"company": "HealthBridge Agency", "title": "Health Insurance Call Center Rep", "location": "Tampa, FL", "source": "Craigslist", "link": "https://tampa.craigslist.org", "search_term": "insurance call center", "description": "Busy health insurance call center seeking experienced reps. Inbound call volume guaranteed."},
    {"company": "Summit Benefits Group", "title": "Final Expense Agent – Warm Transfers Available", "location": "Denver, CO", "source": "ZipRecruiter", "link": "https://www.ziprecruiter.com", "search_term": "live transfer final expense", "description": "We specialize in live transfer final expense sales. Agents can expect 5–10 live transfers per day."},
    {"company": "Guardian Medicare Solutions", "title": "Medicare Supplement Specialist (Inbound Leads)", "location": "Las Vegas, NV", "source": "Glassdoor", "link": "https://www.glassdoor.com", "search_term": "warm transfers medicare", "description": "Real-time warm transfers from compliant marketing campaigns. No prospecting."},
    {"company": "Pacific Health Insurance Agency", "title": "ACA / Obamacare Enrollment Agent", "location": "Los Angeles, CA", "source": "Indeed", "link": "https://www.indeed.com", "search_term": "ACA health insurance", "description": "Seasonal ACA enrollment agents needed. Inbound leads provided from our online campaigns."},
    {"company": "American Senior Benefits", "title": "Life Insurance Sales – Home Office", "location": "Remote", "source": "ZipRecruiter", "link": "https://www.ziprecruiter.com", "search_term": "life insurance agent", "description": "Work from home with pre-qualified leads. Final expense and guaranteed issue life insurance."},
    {"company": "Capital City Insurance", "title": "Health Sales Representative", "location": "Austin, TX", "source": "Craigslist", "link": "https://austin.craigslist.org", "search_term": "health insurance sales", "description": "Full-time health insurance sales. Leads are sourced and provided. Commission-only position."},
]


# ─────────────────────────────────────────────
#  MAIN SCRAPER
# ─────────────────────────────────────────────

def run_scrapers() -> list:
    all_leads = []
    scrapers = [
        scrape_indeed,
        scrape_glassdoor,
        scrape_ziprecruiter,
        scrape_craigslist,
    ]
    for term in SEARCH_TERMS:
        print(f"\n🔍 Searching: '{term}'")
        for scraper in scrapers:
            try:
                results = scraper(term)
                all_leads.extend(results)
                print(f"    ✅ {len(results)} results")
            except Exception as e:
                print(f"    ❌ Error: {e}")
    return all_leads

# ─────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────

COL_WIDTHS_LEADS = [30, 40, 20, 15, 45, 30, 12, 20, 28, 35, 18]
COL_WIDTHS_SOURCE = [15, 30, 40, 20, 45, 12]
COL_WIDTHS_TRACKER = [30, 20, 28, 18, 18, 22, 22]

def style_header_row(ws, col_count: int):
    h = hdr_style()
    for col in range(1, col_count + 1):
        apply_style(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 28

def write_leads_tab(ws, leads: list):
    headers = [
        "Company", "Job Title", "Location", "Source", "Job Link",
        "Search Term", "Priority", "Contact Name", "Contact Email",
        "Notes", "Status",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for r_idx, lead in enumerate(leads, start=2):
        priority = lead.get("priority", "Normal")
        is_high  = priority == "HIGH"

        row_data = [
            lead.get("company", ""),
            lead.get("title", ""),
            lead.get("location", ""),
            lead.get("source", ""),
            lead.get("link", ""),
            lead.get("search_term", ""),
            priority,
            "",   # Contact Name
            "",   # Contact Email
            "",   # Notes
            "Not Contacted",
        ]
        ws.append(row_data)

        # Link cell
        link_cell = ws.cell(row=r_idx, column=5)
        link_url  = lead.get("link", "")
        if link_url:
            link_cell.hyperlink = link_url
            link_cell.font = Font(color="0563C1", underline="single", name="Calibri", size=10)

        # Row fill
        if is_high:
            fill = PatternFill("solid", fgColor=GREEN_FILL)
            font_color = GREEN_FONT
        else:
            fill = PatternFill("solid", fgColor=ROW_ALT if r_idx % 2 == 0 else ROW_WHITE)
            font_color = "1A2D42"

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if col != 5:  # keep hyperlink font on link column
                cell.font = Font(
                    name="Calibri", size=10, color=font_color,
                    bold=(is_high and col == 7)
                )

    # Column widths
    for i, w in enumerate(COL_WIDTHS_LEADS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def write_source_tab(ws, leads: list):
    headers = ["Source", "Company", "Job Title", "Location", "Job Link", "Priority"]
    ws.append(headers)
    style_header_row(ws, len(headers))

    sources_order = ["Indeed", "Glassdoor", "ZipRecruiter", "Craigslist"]
    sorted_leads  = sorted(leads, key=lambda l: (sources_order.index(l["source"]) if l["source"] in sources_order else 9))

    for r_idx, lead in enumerate(sorted_leads, start=2):
        is_high = lead.get("priority") == "HIGH"
        ws.append([
            lead.get("source",""),
            lead.get("company",""),
            lead.get("title",""),
            lead.get("location",""),
            lead.get("link",""),
            lead.get("priority","Normal"),
        ])

        link_cell = ws.cell(row=r_idx, column=5)
        link_url  = lead.get("link","")
        if link_url:
            link_cell.hyperlink = link_url
            link_cell.font = Font(color="0563C1", underline="single", name="Calibri", size=10)

        fill = PatternFill("solid", fgColor=GREEN_FILL if is_high else (ROW_ALT if r_idx % 2 == 0 else ROW_WHITE))
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col != 5:
                cell.font = Font(name="Calibri", size=10, color="1A2D42")

    for i, w in enumerate(COL_WIDTHS_SOURCE, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def write_tracker_tab(ws, leads: list):
    headers = [
        "Company", "Contact Name", "Email",
        "Date Reached Out", "Follow Up Date", "Response", "Outcome",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for r_idx, lead in enumerate(leads, start=2):
        ws.append([
            lead.get("company",""),
            "",  # Contact Name
            "",  # Email
            "",  # Date Reached Out
            "",  # Follow Up Date
            "",  # Response
            "",  # Outcome
        ])
        fill = PatternFill("solid", fgColor=ROW_ALT if r_idx % 2 == 0 else ROW_WHITE)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(name="Calibri", size=10, color="1A2D42")

    for i, w in enumerate(COL_WIDTHS_TRACKER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def build_excel(leads: list, filename: str):
    wb = openpyxl.Workbook()

    # ── Tab 1: Insurance Leads ──
    ws1 = wb.active
    ws1.title = "Insurance Leads"

    high  = [l for l in leads if l.get("priority") == "HIGH"]
    normal = [l for l in leads if l.get("priority") != "HIGH"]
    sorted_leads = high + normal

    write_leads_tab(ws1, sorted_leads)
    print(f"\n📊 Tab 1 written: {len(sorted_leads)} leads ({len(high)} HIGH PRIORITY)")

    # ── Tab 2: By Source ──
    ws2 = wb.create_sheet("By Source")
    write_source_tab(ws2, leads)
    print("📊 Tab 2 written: By Source")

    # ── Tab 3: Outreach Tracker ──
    ws3 = wb.create_sheet("Outreach Tracker")
    write_tracker_tab(ws3, sorted_leads)
    print("📊 Tab 3 written: Outreach Tracker")

    wb.save(filename)


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  CallBound Media – Insurance Agency Lead Finder")
    print("=" * 60)

    print("\n🚀 Starting scraper...\n")
    leads = run_scrapers()

    if not leads:
        print("\n⚠ No live results gathered. Using sample data instead.")
        leads = SAMPLE_DATA

    # Assign priority
    for lead in leads:
        lead["priority"] = get_priority(lead)

    # Deduplicate
    before = len(leads)
    leads  = dedup(leads)
    print(f"\n🔗 Deduplicated: {before} → {len(leads)} unique leads")

    high_count = sum(1 for l in leads if l.get("priority") == "HIGH")
    print(f"⭐ HIGH PRIORITY leads: {high_count}")

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file  = f"insurance_leads_{timestamp}.xlsx"
    out_path  = os.path.join(os.path.dirname(os.path.abspath(__file__)), out_file)

    print(f"\n💾 Building Excel file: {out_file}")
    build_excel(leads, out_path)

    print(f"\n✅ Done! File saved to:")
    print(f"   {out_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
